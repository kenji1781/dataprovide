from django.shortcuts import render
from django.http import HttpResponse
from .models import machine_data #データベースモデル
from .forms import dateForm, main_appForm #フォーム
from django.core.paginator import Paginator,EmptyPage,PageNotAnInteger #ページ制御
import datetime
from openpyxl import * #workbook,load_workbook


def index(request, num=1):
    find = []
    datechk = []
   
    #form = main_appForm()
    datechk = machine_data.objects.all() #.order_by('machine_name','unit_no','-date_y','date_m','date_d')
    
    for chk in datechk:
        if chk.date_ymd == None:
            year_i  = int(chk.date_y)
            month_i = int(chk.date_m)
            day_i = int(chk.date_d)
            chk.date_ymd = datetime.date(year_i,month_i,day_i)
            chk.save()
    
    find = machine_data.objects.all().order_by('-date_ymd','machine_name','unit_no','course_no')
    #find = machine_data.objects.all().order_by('machine_name','unit_no','-date_y','date_m','date_d')
    paginator = Paginator(find, 7)
      

    try:
        data = paginator.get_page(num)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(1)
        
    params = {
        'title': 'DataProvideSystem',
        'msg':'データプロバイドシステム',
        'form': dateForm(),
        'data': data,
        }

    return render(request, 'main_app/index.html', params)

#****************************************************************
def find(request, num=1):
    data = []
    
    if(request.method != 'GET'):

        if(request.method == 'POST'):
            find = []
            global find_data
            #obj = machine_data()
            #input_data = dateForm(request.POST,instance=obj)
            m_n = request.POST['machine_name']
            u_n = request.POST["unit_no"]
            d_m = request.POST["input_date"]
            
            unit_req = u_n.split() #ユニット№書いた分だけ検索

            d_m = datetime.datetime.strptime(d_m,'%Y-%m-%d') #datetime型に変換
            #****************7日分日付データ保管****************
            date_6 = d_m - datetime.timedelta(days=6)
            date_5 = d_m - datetime.timedelta(days=5)
            date_4 = d_m - datetime.timedelta(days=4)
            date_3 = d_m - datetime.timedelta(days=3)
            date_2 = d_m - datetime.timedelta(days=2)
            date_1 = d_m - datetime.timedelta(days=1)
            date_0 = d_m

            start_date = date_6 #検索用に変数用意
            #****************date型に変換****************
            d_m = datetime.datetime.date(d_m)
            start_date = datetime.datetime.date(start_date)
            date_6 = datetime.datetime.date(date_6)
            date_5 = datetime.datetime.date(date_5)
            date_4 = datetime.datetime.date(date_4)
            date_3 = datetime.datetime.date(date_3)
            date_2 = datetime.datetime.date(date_2)
            date_1 = datetime.datetime.date(date_1)
            date_0 = datetime.datetime.date(date_0)

             #****************dateのstr型用意****************

            d_m = d_m.strftime('%Y-%m-%d') #str型に変換
            start_date = start_date.strftime('%Y-%m-%d') #str型に変換
            


            find = machine_data.objects.filter(machine_name=m_n)\
                                        .filter(unit_no__in=unit_req)\
                                        .filter(date_ymd__range=[start_date,d_m])\
                                        .order_by('machine_name','unit_no','-date_ymd')
                                      
            find_data = find #グローバル変数にデータを渡す
                                        
            unit_req = [int(s) for s in unit_req]
        
            
            wb = load_workbook('./DS_4JDC001Format.xlsx')
            
            for j in find: #データ分回す
                print(type(j.date_ymd))
                print(j.date_ymd)
                for i in unit_req: #設定号機分回す
                    for sh in wb.sheetnames: #シート名読出
                        sh_i = int(sh)
                        if i == sh_i and sh == j.unit_no: #データの号機
                            
                            if date_0 == j.date_ymd:
                                #*********日付書込*********
                                date_ind = j.date_ymd
                                date_ind = date_ind.strftime('%Y-%m-%d')
                                ws = wb[j.unit_no]
                                ws.cell(5,4).value = date_ind
                                print(j.course_no)
                                print(type(j.course_no))
                                for k in range(16):
                                    if j.course_no == k:
                                        r = k+6
                                        ws.cell(r,2).value = j.drying_time
                                        ws.cell(r,3).value = j.run_count
                                        ws.cell(r,4).value = j.run_time_m * 60 + j.run_time_s
                                        ws.cell(r,5).value = j.gas_usage

                                    
                            elif date_1 == j.date_ymd:
                                #*********日付書込*********
                                date_ind = j.date_ymd
                                date_ind = date_ind.strftime('%Y-%m-%d')
                                ws = wb[j.unit_no]
                                ws.cell(5,7).value = date_ind
                                print(j.course_no)
                                print(type(j.course_no))
                                for k in range(16):
                                    if j.course_no == k:
                                        r = k+6
                                            #ws.cell(r,2).value = j.drying_time
                                        ws.cell(r,6).value = j.run_count
                                        ws.cell(r,7).value = j.run_time_m * 60 + j.run_time_s
                                        ws.cell(r,8).value = j.gas_usage
                            
                            elif date_2 == j.date_ymd:
                                #*********日付書込*********
                                date_ind = j.date_ymd
                                date_ind = date_ind.strftime('%Y-%m-%d')
                                ws = wb[j.unit_no]
                                ws.cell(5,10).value = date_ind
                                print(j.course_no)
                                print(type(j.course_no))
                                for k in range(16):
                                    if j.course_no == k:
                                        r = k+6
                                            #ws.cell(r,2).value = j.drying_time
                                        ws.cell(r,9).value = j.run_count
                                        ws.cell(r,10).value = j.run_time_m * 60 + j.run_time_s
                                        ws.cell(r,11).value = j.gas_usage
                                  
                            elif date_3 == j.date_ymd:
                                #*********日付書込*********
                                date_ind = j.date_ymd
                                date_ind = date_ind.strftime('%Y-%m-%d')
                                ws = wb[j.unit_no]
                                ws.cell(5,13).value = date_ind
                                print(j.course_no)
                                print(type(j.course_no))
                                for k in range(16):
                                    if j.course_no == k:
                                        r = k+6
                                            #ws.cell(r,2).value = j.drying_time
                                        ws.cell(r,12).value = j.run_count
                                        ws.cell(r,13).value = j.run_time_m * 60 + j.run_time_s
                                        ws.cell(r,14).value = j.gas_usage    
                        
                            elif date_4 == j.date_ymd:
                                #*********日付書込*********
                                date_ind = j.date_ymd
                                date_ind = date_ind.strftime('%Y-%m-%d')
                                ws = wb[j.unit_no]
                                ws.cell(5,16).value = date_ind
                                print(j.course_no)
                                print(type(j.course_no))
                                for k in range(16):
                                    if j.course_no == k:
                                        r = k+6
                                            #ws.cell(r,2).value = j.drying_time
                                        ws.cell(r,15).value = j.run_count
                                        ws.cell(r,16).value = j.run_time_m * 60 + j.run_time_s
                                        ws.cell(r,17).value = j.gas_usage    
                        
                            elif date_5 == j.date_ymd:
                                #*********日付書込*********
                                date_ind = j.date_ymd
                                date_ind = date_ind.strftime('%Y-%m-%d')
                                ws = wb[j.unit_no]
                                ws.cell(5,19).value = date_ind
                                print(j.course_no)
                                print(type(j.course_no))
                                for k in range(16):
                                    if j.course_no == k:
                                        r = k+6
                                            #ws.cell(r,2).value = j.drying_time
                                        ws.cell(r,18).value = j.run_count
                                        ws.cell(r,19).value = j.run_time_m * 60 + j.run_time_s
                                        ws.cell(r,20).value = j.gas_usage    
                        
                            elif date_6 == j.date_ymd:
                                #*********日付書込*********
                                date_ind = j.date_ymd
                                date_ind = date_ind.strftime('%Y-%m-%d')
                                ws = wb[j.unit_no]
                                ws.cell(5,22).value = date_ind
                                print(j.course_no)
                                print(type(j.course_no))
                                for k in range(16):
                                    if j.course_no == k:
                                        r = k+6
                                            #ws.cell(r,2).value = j.drying_time
                                        ws.cell(r,21).value = j.run_count
                                        ws.cell(r,22).value = j.run_time_m * 60 + j.run_time_s
                                        ws.cell(r,23).value = j.gas_usage
                        
                        
 
                        
                        
                        #elif i == 102:
            
            wb.save('./test1.xlsx')                                  


                            
    
     #**********ページ制御**********
    
    paginator = Paginator(find_data, 7)

    data = paginator.get_page(num)
   
    params = {
        'title': 'DataProvideSystem',
        'msg':'データプロバイドシステム',
        'form': dateForm(),
        'data': data,
        }

    return render(request, 'main_app/find.html', params)

#****************************************************************
def date(request, num=1):
    find = []
    data = []
   
    if(request.method == 'POST'):
        form = main_appForm(request.POST) #入力した内容表示の為
        #**********フォームmachine**********
        m = request.POST['machi']
        #**********フォームunit**********
        u = request.POST['unit']
        #**********フォームcourse**********
        #course = request.POST['course']
        #**********フォームyear検索**********
        #year = request.POST['year']
        #**********フォームmonth検索**********
        #month = request.POST['month']
        #**********フォームday検索**********
        #day = request.POST['day']
        find = machine_data.objects.filter(machine_name=m,unit_no=u)\
                                    .order_by('machine_name','unit_no','-date_y','date_m','date_d')
        #**********ページ制御**********
    else:
        form = main_appForm()
        find = machine_data.objects.all().order_by('machine_name','unit_no','-date_y','date_m','date_d')
    paginator = Paginator(find, 7)

    try:
        data = paginator.page(num)
        for i in data:
            year  = int(i.date_y)
            month = int(i.date_m)
            day = int(i.date_d)
            date = datetime.date(year,month,day)

            if date != i.date_ymd:
                i.date_ymd = date
                i.save() 


    except PageNotAnInteger:
        data = paginator.page(1)
        for i in data:
            year  = int(i.date_y)
            month = int(i.date_m)
            day = int(i.date_d)
            date = datetime.date(year,month,day)

            if date != i.date_ymd:
                i.date_ymd = date
                i.save() 

    except EmptyPage:
        data = paginator.page(1)
        for i in data:
            year  = int(i.date_y)
            month = int(i.date_m)
            day = int(i.date_d)
            date = datetime.date(year,month,day)

            if date != i.date_ymd:
                i.date_ymd = date
                i.save() 
    
    params = {
        'title': 'DataProvideSystem',
        'msg':'データプロバイドシステム',
        'form': dateForm(),
        'data': data,
        }

    return render(request, 'main_app/find.html', params)
# Create your views here.
